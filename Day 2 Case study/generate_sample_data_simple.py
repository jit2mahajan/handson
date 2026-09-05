#!/usr/bin/env python3
"""
Simple sample data generator using only openpyxl (no pandas dependency).
"""

import random
from openpyxl import Workbook

def main():
    """Generate and save sample pharmaceutical shipment data."""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"
    
    # Add headers
    headers = [
        'shipment_id', 'origin', 'destination', 'avg_temperature', 'avg_humidity',
        'transit_days', 'handling_violations', 'temp_sensitive', 'temperature_excursion', 'improper_storage'
    ]
    ws.append(headers)
    
    # Sample data sources
    origins = ['NYC Hub', 'LA Facility', 'Chicago Center', 'Houston Depot', 'Atlanta Terminal']
    destinations = [
        'Boston Hospital', 'Miami Clinic', 'Seattle Medical', 'Denver Health',
        'Dallas Center', 'San Francisco Pharma', 'Phoenix Med Center', 'Austin Health'
    ]
    
    # Set seed for reproducibility
    random.seed(42)
    
    # Generate 50 sample shipments
    print("Generating 50 sample pharmaceutical shipments...")
    for i in range(1, 51):
        row_data = [
            f'SHP-{i:05d}',
            random.choice(origins),
            random.choice(destinations),
            round(random.uniform(8, 32), 1),  # avg_temperature
            round(random.uniform(30, 85), 1),  # avg_humidity
            random.randint(3, 24),  # transit_days
            random.randint(0, 4),   # handling_violations
            random.choice([True, False]),  # temp_sensitive
            random.choice([True, False]),  # temperature_excursion
            random.choice([True, False]),  # improper_storage
        ]
        ws.append(row_data)
    
    # Add some intentional high-risk shipments for demo
    ws['D2'] = 32        # High temperature
    ws['E3'] = 85        # High humidity
    ws['F4'] = 21        # Long transit
    ws['G5'] = 4         # Many violations
    ws['I6'] = True      # Temperature excursion
    ws['H7'] = True      # Temp sensitive
    ws['J7'] = True      # Improper storage
    
    # Save file
    output_path = 'sample_data.xlsx'
    wb.save(output_path)
    
    print(f"\n✓ Sample data file created: {output_path}")
    print(f"✓ Contains 50 sample pharma shipments with realistic risk profiles")
    print("\nYou can now run the app with:")
    print("  streamlit run app.py")

if __name__ == '__main__':
    main()
